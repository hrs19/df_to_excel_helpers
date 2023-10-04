import os
import xlwings as xw
from openpyxl import load_workbook

class ExcelHandler:
    def __init__(self, file_path, library='openpyxl'):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file '{file_path}' does not exist.")
        self.file_path = file_path
        self.library = library.lower()

        if self.library not in ['openpyxl', 'xlwings']:
            raise ValueError("Invalid library choice. Use 'openpyxl' or 'xlwings'.")

    def read_excel(self, sheet_name):
        if self.library == 'openpyxl':
            return self.read_excel_openpyxl(sheet_name)
        elif self.library == 'xlwings':
            return self.read_excel_xlwings(sheet_name)

    def read_excel_openpyxl(self, sheet_name):
        if not self.file_path.endswith(".xlsx"):
            raise ValueError("Unsupported Excel file format. Please use .xlsx.")

        # Use openpyxl to read Excel
        wb = load_workbook(self.file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        wb.close()
        return data

    def read_excel_xlwings(self, sheet_name):
        if not self.file_path.endswith(".xlsx"):
            raise ValueError("Unsupported Excel file format. Please use .xlsx.")

        # Use xlwings to read Excel
        app = xw.App(visible=False)
        wb = app.books.open(self.file_path)
        if sheet_name not in wb.sheets:
            wb.close()
            app.quit()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        ws = wb.sheets[sheet_name]
        data = ws.used_range.options(ndim=2).value
        wb.close()
        app.quit()
        return data

    def update_excel(self, sheet_name, data):
        if self.library == 'openpyxl':
            self.update_excel_openpyxl(sheet_name, data)
        elif self.library == 'xlwings':
            self.update_excel_xlwings(sheet_name, data)

    def update_excel_openpyxl(self, sheet_name, data):
        if not self.file_path.endswith(".xlsx"):
            raise ValueError("Unsupported Excel file format. Please use .xlsx.")

        if data is None or not isinstance(data, list):
            raise ValueError("Data parameter must be a list of lists for the update operation.")

        # Use openpyxl to update Excel
        wb = load_workbook(self.file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        ws = wb[sheet_name]
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_value)
        wb.save(self.file_path)
        wb.close()

    def update_excel_xlwings(self, sheet_name, data):
        if not self.file_path.endswith(".xlsx"):
            raise ValueError("Unsupported Excel file format. Please use .xlsx.")

        if data is None or not isinstance(data, list):
            raise ValueError("Data parameter must be a list of lists for the update operation.")

        # Use xlwings to update Excel
        app = xw.App(visible=False)
        wb = app.books.open(self.file_path)
        if sheet_name not in wb.sheets:
            wb.close()
            app.quit()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        ws = wb.sheets[sheet_name]
        ws.range("A1").value = data
        wb.save()
        wb.close()
        app.quit()

"""
# Example usage:
# Create an ExcelHandler instance for the Excel file, specifying the library to use
excel_handler_openpyxl = ExcelHandler("example.xlsx", library='openpyxl')
excel_handler_xlwings = ExcelHandler("example.xlsx", library='xlwings')

# Read data from an Excel file using openpyxl
data_openpyxl = excel_handler_openpyxl.read_excel("Sheet1")
print(data_openpyxl)

# Update data in an Excel file using xlwings
new_data_xlwings = [["Updated", "Data"], [1, 2], [3, 4]]
excel_handler_xlwings.update_excel("Sheet1", new_data_xlwings)
"""